# 정부장님 스타일의 pdf 기획서 (표지에는 페이지 표시 안된 기획서)에선 정상 동작됨

import pdfplumber                       # pdf 제어를 위해
import re                               # 정규식, split 등 사용을 위해
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font
import smtplib                                          # 이메일 사용을 위해
from email.encoders import encode_base64
from email.header import Header
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from email import encoders
import os                                               # 파일 제목 가져오기 위해

path = r"C:\Users\mung9\Documents\RPAStudy\SalesKit_website Storyboard_v0.7.pdf"           # pdf 경로
with pdfplumber.open (path) as pdf:                                                         # with 사용으로 사용 완료후 자동으로 메모리 닫기
    all_text = ""
    for page in pdf.pages:
        page_text = page.extract_text()                                                     # 각 페이지 순회하며 텍스트 내용 누적하기
        all_text += page_text

#print (all_text)
print (len(all_text))
print(type(all_text))

split_page  = re.split(r'[\n]', all_text)       # 개행 문자로 1차 가공

for i in split_page:
    print (i)

str_re_page = str(split_page)
page_num = re.findall (r"Page: \d?\d?\d?", str_re_page) # 정규식 이용하여 페이지 번호 추출하기, Page : nnn 식으로 짤림

num_modified_list = []
for i in page_num:
    temp = i.split('Page: ')                            # Page : nnn 에서 Page : 부분 짜르기
    num_modified_list.append(temp[1].strip())           # 숫자 부분만 새로운 리스트에 담기, strip()로 양쪽의 공백 제거하기

for i in num_modified_list:
    print(i)

#title_page = re.findall (r".*?Screen",str_re_page, re.MULTILINE)
#print (title_page)

modified_list = []
include_screen = re.findall (r".*Screen No. : *", all_text)         # Screen No. : 형식이 포함된 모든 텍스트 찾기
#print(type(include_screen))

for i in include_screen:
     # print(i)
     if 'Screen No. :' in i:                                        # Screen No. : 있다면
         TempList = i.split('Screen No. :')                         # Screen No. : 기준으로 짜르기
         modified_list.append(TempList[0].strip())                  # Screen No. : 기준 앞 부분 텍스트만 새로운 리스트에 담기
     else:
         modified_list.append(i)

for i in modified_list:
    print(i)

print(f"Length of modified_list: {len(modified_list)}")
print (modified_list[0])

excel_path = r"C:\Users\mung9\Documents\RPAStudy\기획서 페이지 정리.xlsx"         # 엑셀 경로
wb = openpyxl.load_workbook(excel_path)                                         # 기존 엑셀 열기
ws = wb.active                                                                  # 활성화 된 시트 선택 (시트 1)

ws["B1"] = os.path.basename(path)                                               # pdf 기획서 제목 입력
ws["B1"].font = Font(size="15", bold=True)

base_title = os.path.basename(path)

base_temp = base_title.split('Storyboard_')
base_temp1 = base_temp[1].split('.pdf')
base_modi = base_temp1[0]

'''base_temp = base_title.split('.pdf')
base_modi = base_temp[0]
print (base_modi)'''

for i in range(len(num_modified_list)):
    if i == 0:
        ws["B"+str(3+i)] = num_modified_list[i]
        ws["B"+str(3+i)].border = Border(top = Side(border_style='thick'),               # 셀 테두리 설정
                                     bottom= Side(border_style='thin'), 
                                     left= Side(border_style='thin'), 
                                     right= Side(border_style='thin'))
        ws["B"+str(3+i)].alignment = Alignment(horizontal= 'center', vertical= 'center')        # 셀 가운데 정렬
    else:
        ws["B"+str(3+i)] = num_modified_list[i]
        ws["B"+str(3+i)].border = Border(top = Side(border_style='thin'),               # 셀 테두리 설정
                                     bottom= Side(border_style='thin'), 
                                     left= Side(border_style='thin'), 
                                     right= Side(border_style='thin'))
        ws["B"+str(3+i)].alignment = Alignment(horizontal= 'center', vertical= 'center')        # 셀 가운데 정렬

#differ_len = len(num_modified_list) - len(modified_list)
j = 0
for i in range(len(num_modified_list)):                                     # 페이지 제목 엑셀에 입력 시작
    if i == 0:  
        ws["C"+str(3+i)] = os.path.basename(path) + ' 의 첫 페이지'                                         # 첫 페이지 일때, 해당 텍스트 입력
        ws["C"+str(3+i)].border = Border(top = Side(border_style='thin'), 
                                        bottom= Side(border_style='thin'), 
                                        left= Side(border_style='thin'), 
                                        right= Side(border_style='thin'))
    elif j == len(modified_list):                                        # i 가 넘버 리스트보다 개수가 1개 작을때, 해당 텍스트 입력
        ws["C"+str(3+i)] = 'End Of Document'
        ws["C"+str(3+i)].border = Border(top = Side(border_style='thin'), 
                                        bottom= Side(border_style='thin'), 
                                        left= Side(border_style='thin'), 
                                        right= Side(border_style='thin'))
    else:
        ws["C"+str(3+i)] = modified_list[j]
        ws["C"+str(3+i)].border = Border(top = Side(border_style='thin'), 
                                        bottom= Side(border_style='thin'), 
                                        left= Side(border_style='thin'), 
                                        right= Side(border_style='thin'))
        j += 1

for i in range(len(num_modified_list)):
    ws["D"+str(3+i)] = base_modi                             # 3열부터 입력 시작, 페이지 넘버 엑셀에 입력
    ws["D"+str(3+i)].border = Border(top = Side(border_style='thin'),               # 셀 테두리 설정
                                     bottom= Side(border_style='thin'), 
                                     left= Side(border_style='thin'), 
                                     right= Side(border_style='thin'))
    ws["D"+str(3+i)].alignment = Alignment(horizontal= 'center', vertical= 'center')        # 셀 가운데 정렬

wb.save(excel_path)

msg = MIMEMultipart()                                                           # 첨부 파일이 포함된 이메일을 보내기 위해 사용

from_addr = 'hwjung@inzisoft.com'                                               # 발신인
#to_addr = ['hwjung@inzisoft.com', 'hrkim@inzisoft.com', 'dhyoo@inzisoft.com']             # 다중 수신인 이용시
to_addr = ['hwjung@inzisoft.com'] 

smtp = smtplib.SMTP('smtp.gmail.com', 587)                                      # 지메일 smtp 주소 및 포트 번호
smtp.starttls()                                                                 # tls 연결 방식 시작

smtp.login('hwjung@inzisoft.com', 'nara0805~!')

subject = '기획서 페이지 번호 및 제목 추출 결과 공유'
body = '''
기획서의 페이지 번호 및 제목 추출 결과 공유 합니다. (★ 각 표지에 'Page :' 텍스트가 존재하지 않는 기획서)
자세한 내용은 첨부된 엑셀 파일 참조 부탁 드립니다.
감사합니다.
'''
msg = MIMEMultipart()                                                       # 파일을 첨부하기 위해
msg['Subject'] = subject
msg['From'] = from_addr
#msg['To'] = to_addr                                                        # 수신인 단일인 일때 사용
msg['To'] = ",".join(to_addr)                                               # 수신인 다인일때 사용
msg.attach(MIMEText(body, 'plain'))

excel_filename = os.path.basename(excel_path)
with open(excel_path, 'rb') as excel_att:
    excel = MIMEApplication(excel_att.read())                                           # 데이터 파일 첨부를 위해
    excel.add_header('Content-Disposition','attachment', filename= excel_filename)
    msg.attach(excel)

imagefile = r'C:\Users\mung9\Documents\RPAStudy\잔망루피_건배.png'

filename = os.path.basename(imagefile)
with open(imagefile, 'rb') as image_att:
    img = MIMEImage(image_att.read())                                                   # 이미지 파일 첨부를 위해
    '''part.set_payload(attachment.read())
    encoders.encode_base64(part)'''
    img.add_header('Content-Disposition','attachment', filename= filename)
    msg.attach(img)

smtp.sendmail (msg['From'], msg['To'], msg.as_string())                    # 수신인 단일인 일때 사용

#smtp.sendmail (msg['From'], to_addr, msg.as_string())                          # 수신인 다인일때 사용
smtp.quit()
