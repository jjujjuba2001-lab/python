import selenium
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

import tkinter as tk                    # 윈도우 에러 메시지 띄우기 위해 import
from tkinter import messagebox

import openpyxl                                 # 엑셀 컨트롤
from openpyxl.styles import Border
from openpyxl.drawing.image import Image            # 엑셀에 이미지 삽입 위해
from PIL import Image as PILImage

id = ['hwjung@inzisoft.com']                  # 테스트 계정 id
pw = ['test1234']                           # 테스트 계정 pw

'''driver = webdriver.Chrome()
Options = Options()
Options.add_experimental_option("detach", True)'''

# 드라이버 설정 수정
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()  # 옵션 추가 가능

driver = webdriver.Chrome(service=service, options=options)

def get_col_width_row_height(img_width, img_height):                    # 이미지 크기에 따라 셀 높이, 너비 조정 항수
    col_width = img_width*63.2/504.19
    row_height = img_height*225.35/298.96
    return (col_width, row_height)

excel_path = r"C:\Users\mung9\Downloads\QA 일일 업무 진행 현황.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)                 # data_only 옵션은 수식으로 계산된 값 사용할때 사용, False 시 수식 자체 가져옴
ws = wb['2025 일일 업무 진행 현황']                                        # 특정 시트로 파일 열기

total_rows = ws.max_row
print ("엑셀 전체 행 개수는 " + str(total_rows) + " 입니다.")

target_value = None              # 전역 변수 선언

root = tk.Tk()
var = tk.IntVar()

def check_selection():
    global target_value             # 전역 변수 사용 선언
    target_value = int(var.get())  # 입력값 가져오기
    print (target_value)
    root.destroy()

root.title("TBMS 에 입력할 월을 선택하세요.")
root.geometry("350x300")  # 창 크기 설정 (너비 x 높이)

# 왼쪽과 오른쪽을 나눌 Frame 생성
frame_left = tk.Frame(root, relief="solid", borderwidth=1)  # 선 추가
frame_right = tk.Frame(root, relief="solid", borderwidth=1)  # 선 추가

frame_left.pack(side="left", expand=True, padx=3, pady=7)
frame_right.pack(side="right", expand=True, padx=3, pady=7)

# 선택 변수
var = tk.IntVar()

# 왼쪽에 배치할 버튼 (1~6)
for i in range(1, 7):
    radio = tk.Radiobutton(frame_left, text=f"{i}월", variable=var, value=i)
    radio.pack(anchor="e", pady=3,padx=50)  # 왼쪽 정렬

# 오른쪽에 배치할 버튼 (7~12)
for i in range(7, 13):
    radio = tk.Radiobutton(frame_right, text=f"{i}월", variable=var, value=i)
    radio.pack(anchor="w",pady=3,padx=50)  # 왼쪽 정렬

button = tk.Button(root, text="확 인", width=30, height=2, relief="ridge", borderwidth=2,
                   highlightbackground="blue", highlightcolor="red", bg="sky blue", command=check_selection)  # 버튼 생성
button.place(relx=0.5, rely=1, anchor="s", y=-10) 

root.mainloop()
#-------------------------------------------------------------------------------------------------------------------------

# 찾고 싶은 값과 열 지정
#target_value = 2  # 여기에 찾고 싶은 값 (월) 입력
target_column = "B"  # 찾고 싶은 열 지정 (예: "B")

# 전체 행 개수 확인
total_rows = ws.max_row

# 특정 값이 처음 나타나는 행 찾기
first_occurrence_row = None  # 첫 번째로 찾은 행 번호 저장용
b_values = []               # 엑셀 B열 모든 값 담을 리스트

for row in range(1, total_rows + 1):  # 1부터 마지막 행까지 반복
    cell_value = ws[f"{target_column}{row}"].value  # 특정 열의 값 가져오기
    if cell_value is not None:         # 빈칸(None) 제외
        b_values.append(cell_value)    # 리스트에 추가

count_month = b_values.count(target_value)                      # 타겟 월이 엑셀 B열에 몇개나 있나 확인하기
print(f"B열에서 {target_value} 값은 총 {count_month}개 있습니다.")

for row in range(1, total_rows + 1):  # 1부터 마지막 행까지 반복
    cell_value = ws[f"{target_column}{row}"].value  # 특정 열의 값 가져오기
    if cell_value == target_value:
        first_occurrence_row = row
        break  # 첫 번째로 찾았으니 반복 종료

# 결과 출력
if first_occurrence_row:
    print(f"'{target_value}' 값이 처음 나타나는 행: {first_occurrence_row}")
else:
    print(f"'{target_value}' 값이 해당 열에 없습니다.")

#-------------------------------------------------------------------------------------------------

driver.get("https://tbms.mobileleader.com:8010/login")                 # TBMS 사이트 접속
driver.maximize_window()
#Options.add_experimental_option("excludeSwitches", ["enable-automation"])     # 자동화 제어 알림 줄 삭제 , 하지만 안됨
driver.implicitly_wait(5)
print (driver.current_window_handle)

driver.find_element (By.ID, 'plainUserId').send_keys(id)            # id 입력
driver.implicitly_wait(5)
driver.find_element (By.ID, 'plainUserPw').send_keys(pw)            # pw 입력
driver.implicitly_wait(5)
driver.find_element (By.LINK_TEXT, '로그인').click()                 # 로그인 클릭
driver.implicitly_wait(5)

driver.find_element (By.LINK_TEXT, '업무(일반)').click()                # 업무 탭 클릭
driver.implicitly_wait(5)

driver.find_element (By.ID, 'workDate').click()                     # 웹 데이트 피커의 날짜 수집을 위한 클릭
days = driver.find_elements (By.CLASS_NAME, 'ui-state-default')     # 데이트 피커의 날짜 수집
web_day_list = []                                                   # 웹에서 수집한 날짜 담을 빈 리스트 생성

for day in days:                                                    # 웹에서 수집한 날짜 출력을 위한 포 문
    web_day = day.text                                              # days 에서 수집한 날짜 값으로 변환해 변수에 담기
    web_day_list.append(web_day)                                    # 빈 리스트에 하나씩 추가하기

print (web_day_list)

#-------------------------------------------------------------------------------------------------

excel_person_time_row = first_occurrence_row + 2                    # (정현우 기준) 해당월이 처음 시작하는 행 + 2 해야함. 
print (excel_person_time_row)

work_sum = 0                                                        # 해당 월의 등록 업무 총 시간, 초기화

for i in range (count_month):                                 # 엑셀 B열에서 수집한 날짜 중 타겟 월의 개수 만큼 반복하는 포 문
    excel_person_time = ws["I"+ str(excel_person_time_row)].value           # (정현우 기준) 처음 시작은 first_occurrence_row + 2 셀 부터 시작, 엑셀에서 개인 업무 시간 값 가져오기
    print(excel_person_time)
   
    if excel_person_time == None:                         # 엑셀 개인 업무 시간 값이 공란이면  
        excel_person_time_row += 3                        # 다음 날짜의 행으로 이동, (정현우 기준) excel_person_time_row + 3 해야함 
        first_occurrence_row += 3                         # 다음 날짜의 행으로 이동, (정현우 기준) first_occurrence_row (일자, 계약명 등 행 이동 위해) + 3 해야함 
        continue                                          # 아래 내용은 진행하지 않고, 다시 포 문 으로 복귀  
    
    execl_raw_date = ws["C"+ str(first_occurrence_row)].value           # 엑셀 특정 셀 값 (날짜) 변수에 대입. 특정 월이 처음으로 나타나는 행 기준
    execl_raw_date = execl_raw_date.strftime('%Y-%m-%d')     # 날짜 형식을 특정 형식의 텍스트로 변경 
    print (execl_raw_date)

    excel_temp_date = execl_raw_date.split ('-')                # - 로 텍스트로 자르기
    excel_final_date = excel_temp_date[2]                       # 날짜만 변수에 담기 (01, 02 이런식으로 담김)
    excel_final_date = int(excel_final_date)                    # xpath 찾기 위해, 앞에 0 제거 위해 다시 숫자 형식으로 변환
    print (excel_final_date)

    driver.find_element (By.ID, 'workDate').click()             # 데이트 피커 클릭
    driver.implicitly_wait(5)
    try:
        day_click = driver.find_element(By.XPATH, f"//a[@class='ui-state-default' and text()='{excel_final_date}']")      # ※ and 를 사용해 xpath 와 text 를 동시에 만족하는 엘리멘트 찾음. 엑셀 파일에서 수집한 날짜를 데이트 피커에서 선택
        time.sleep(1)
        day_click.click()
    except Exception as e:
        print(f"버튼을 찾는 중 오류 발생: {e}")
        # 등록할 날짜에 업무가 2개 이상이며, 기등록 되어있는 경우 ui-state-default.ui-state-active 를 and 로 분리 및 연결, 3개의 and 모두 만족하는 엘리멘트 찾기
        try:
            activeday_click = driver.find_element(By.XPATH, f"//a[contains(@class, 'ui-state-default') and contains(@class, 'ui-state-active') and text()='{excel_final_date}']") 
            activeday_click.click()
        except Exception as e:
            print(f"활성화된 날짜에서 버튼을 찾는 중 오류 발생: {e}")
            # 등록할 날짜가 오늘이라면 ui-state-default.ui-state-highlight 를 and 로 분리 및 연결, 3개의 and 모두 만족하는 엘리멘트 찾기
            today_click = driver.find_element(By.XPATH, f"//a[contains(@class, 'ui-state-default') and contains(@class, 'ui-state-highlight') and text()='{excel_final_date}']")   
            today_click.click()

    works = driver.find_elements (By.CLASS_NAME, 'main_select01.mr6.wrkSel.prjWrk')     # 웹에서 콤보 박스 내용 수집
    for work in works:                                                    # 수집한 목록 표시
        print (work.text)

    excel_work = ws["F"+ str(first_occurrence_row)].value                     # 엑셀에서 tbms 계약명 값 수집,  특정 월이 처음으로 나타나는 행 기준
    print (excel_work)

    works = driver.find_element (By.CLASS_NAME, 'main_select01.mr6.wrkSel.prjWrk')         # select 사용하기 위해 콤보박스 지정
    select = Select(works)
    select.select_by_visible_text(f"{excel_work}")                                         # 엑셀에서 수집한 tbms 계약명 웹의 콤보박스에서 선택 하기

    times = driver.find_element (By.NAME, 'wrkHour')                                   # tbms 시간 콤보 박스 지정
    print (times.text)
    select = Select(times)

    if excel_person_time == "2시간":                                                   # tbms 시간 콤보 박스 입력 위해, 엑셀에서 수집한 시간과 비교
        select.select_by_value("2")                                                    # value 값으로 선택
        work_sum += 0.0125
    elif excel_person_time == "4시간":
        select.select_by_value("4")
        work_sum += 0.025
    elif excel_person_time == "6시간":
        select.select_by_value("6")
        work_sum += 0.0375
    elif excel_person_time == "8시간":
        select.select_by_value("8")
        work_sum += 0.05

    excel_contents = ws["G" + str(first_occurrence_row)].value         # 엑셀에서 업무 진행 내용 수집, 특정 월이 처음으로 나타나는 행 기준
    print (excel_contents)

    excel_person_time_row += 3                                         # 다음 날짜의 행으로 이동, (정현우 기준) excel_person_time_row + 3 해야함
    first_occurrence_row += 3                                          # 다음 날짜의 행으로 이동, (정현우 기준) first_occurrence_row (일자, 계약명 등 행 이동 위해) + 3 해야함

    driver.find_element (By.CLASS_NAME, 'main_textarea01.mt10.mr10').send_keys (excel_contents)     # tbms 에 업무 진행 내용 입력

    driver.find_element (By.LINK_TEXT, '등록').click()
    time.sleep(1)
    alert = driver.switch_to.alert                          # 크롬 알럿창 제어를 위해
    print(alert.text)
    alert.accept()                                          # 크롬 알럿창의 확인 선택
    time.sleep(1)
    driver.find_element (By.CLASS_NAME, 'main_textarea01.mt10.mr10').clear()                         # tbms 에 입력한 업무 내용 클리어

messagebox.showinfo("알림", f"TBMS 업무 입력 완료\n업무 등록 시간 합계 : {work_sum}")